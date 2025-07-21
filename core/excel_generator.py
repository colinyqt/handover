import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path

class ExcelGenerator:
    def __init__(self):
        self.workbook = None
    
    def generate_compliance_report(self, output_file, data):
        """Generate Excel compliance report - method name matches prompt engine call"""
        
        print(f"📊 Generating Excel report: {output_file}")
        
        # FIX: Validate and fix data structure
        data = self.validate_and_fix_data_structure(data)
        
        # Create workbook
        self.workbook = Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        # Generate sheets
        self.create_summary_sheet(data.get('summary_sheet', {}))
        self.create_compliance_matrix_sheet(data.get('compliance_matrix', {}))
        self.create_meter_specs_sheet(data.get('meter_specs', {}))
        
        # Save file
        self.workbook.save(output_file)
        print(f"✅ Excel file saved: {output_file}")
        
        return True
    
    def validate_and_fix_data_structure(self, data):
        """Validate and fix malformed data structure"""
        
        print(f"🔍 Validating data structure...")
        print(f"Data type: {type(data)}")
        
        # Handle string data (failed JSON parsing)
        if isinstance(data, str):
            print("⚠️ Data is string, attempting to parse as JSON...")
            try:
                data = json.loads(data)
            except json.JSONDecodeError:
                print("❌ Failed to parse string as JSON, using fallback")
                return self.create_fallback_structure()
        
        # Handle non-dict data
        if not isinstance(data, dict):
            print(f"❌ Data is not dict (type: {type(data)}), using fallback")
            return self.create_fallback_structure()
        
        # Check for error in data
        if 'error' in data:
            print(f"❌ Error in data: {data['error']}, using fallback")
            return self.create_fallback_structure()
        
        # Check for required sections
        required_sections = ['summary_sheet', 'compliance_matrix', 'meter_specs']
        missing_sections = [s for s in required_sections if s not in data]
        
        if missing_sections:
            print(f"⚠️ Missing sections: {missing_sections}")
            
            # Try to fix missing sections
            if len(missing_sections) == len(required_sections):
                # All sections missing - complete fallback
                print("❌ All sections missing, using complete fallback")
                return self.create_fallback_structure()
            else:
                # Some sections missing - fix individually
                print("🔧 Fixing missing sections...")
                data = self.fix_missing_sections(data, missing_sections)
        
        print("✅ Data structure validated")
        return data
    
    def fix_missing_sections(self, data, missing_sections):
        """Fix individual missing sections"""
        
        for section in missing_sections:
            if section == 'summary_sheet':
                data['summary_sheet'] = {
                    "title": "Compliance Summary",
                    "data": {
                        "project_name": "Tender Compliance Analysis",
                        "selected_meter": "PM5320 (PM5000 Series)",
                        "analysis_date": "2025-07-02",
                        "generated_by": "colinyqt",
                        "overall_compliance": "Data extraction incomplete",
                        "total_requirements": 0,
                        "status_breakdown": {
                            "fully_compliant": 0,
                            "partially_compliant": 0,
                            "non_compliant": 0
                        }
                    }
                }
            
            elif section == 'compliance_matrix':
                data['compliance_matrix'] = {
                    "title": "Detailed Compliance Matrix",
                    "headers": [
                        "Clause ID", "Category", "Parameter", "Required", 
                        "Meter Spec", "Status", "Justification", "Risk", "Comments"
                    ],
                    "data": [
                        ["N/A", "Error", "Data extraction incomplete", "N/A", "N/A", 
                         "ERROR", "LLM processing incomplete", "High", "Check source analysis file"]
                    ]
                }
            
            elif section == 'meter_specs':
                data['meter_specs'] = {
                    "title": "Selected Meter Specifications",
                    "meter_details": {
                        "model": "Unknown",
                        "series": "Unknown", 
                        "selection_source": "Data extraction incomplete",
                        "specifications": {
                            "status": "Data extraction failed"
                        }
                    }
                }
        
        return data
    
    def create_fallback_structure(self):
        """Create complete fallback structure when all else fails"""
        
        print("🔄 Creating fallback structure...")
        
        return {
            "summary_sheet": {
                "title": "Compliance Summary",
                "data": {
                    "project_name": "Tender Compliance Analysis",
                    "selected_meter": "PM5320 (PM5000 Series)",
                    "analysis_date": "2025-07-02",
                    "generated_by": "colinyqt",
                    "overall_compliance": "Data extraction failed",
                    "total_requirements": 1,
                    "status_breakdown": {
                        "fully_compliant": 0,
                        "partially_compliant": 0,
                        "non_compliant": 1
                    }
                }
            },
            "compliance_matrix": {
                "title": "Detailed Compliance Matrix",
                "headers": [
                    "Clause ID", "Category", "Parameter", "Required", 
                    "Meter Spec", "Status", "Justification", "Risk", "Comments"
                ],
                "data": [
                    ["ERROR", "System", "Data Extraction", "Complete analysis", "Failed", 
                     "NON-COMPLIANT", "LLM processing failed to extract compliance data", 
                     "High", "Check source analysis file format and content"]
                ]
            },
            "meter_specs": {
                "title": "Selected Meter Specifications",
                "meter_details": {
                    "model": "Data extraction failed",
                    "series": "Unknown",
                    "selection_source": "Error - processing failed",
                    "specifications": {
                        "status": "Unable to extract meter specifications",
                        "recommendation": "Check source analysis file and retry"
                    }
                }
            }
        }
    
    def create_summary_sheet(self, summary_data):
        """Create summary sheet"""
        
        ws = self.workbook.create_sheet("Summary")
        
        # Title
        ws['A1'] = summary_data.get('title', 'Compliance Summary')
        ws['A1'].font = Font(size=16, bold=True)
        
        # Data
        data = summary_data.get('data', {})
        
        row = 3
        ws[f'A{row}'] = "Project Name:"
        ws[f'B{row}'] = data.get('project_name', 'Unknown')
        
        row += 1
        ws[f'A{row}'] = "Selected Meter:"
        ws[f'B{row}'] = data.get('selected_meter', 'Unknown')
        
        row += 1
        ws[f'A{row}'] = "Analysis Date:"
        ws[f'B{row}'] = data.get('analysis_date', '2025-07-02')
        
        row += 1
        ws[f'A{row}'] = "Generated By:"
        ws[f'B{row}'] = data.get('generated_by', 'colinyqt')
        
        row += 1
        ws[f'A{row}'] = "Overall Compliance:"
        ws[f'B{row}'] = data.get('overall_compliance', 'Unknown')
        
        row += 1
        ws[f'A{row}'] = "Total Requirements:"
        ws[f'B{row}'] = data.get('total_requirements', 0)
        
        # Status breakdown
        row += 2
        ws[f'A{row}'] = "Compliance Breakdown:"
        ws[f'A{row}'].font = Font(bold=True)
        
        breakdown = data.get('status_breakdown', {})
        
        row += 1
        ws[f'A{row}'] = "Fully Compliant:"
        ws[f'B{row}'] = breakdown.get('fully_compliant', 0)
        
        row += 1
        ws[f'A{row}'] = "Partially Compliant:"
        ws[f'B{row}'] = breakdown.get('partially_compliant', 0)
        
        row += 1
        ws[f'A{row}'] = "Non-Compliant:"
        ws[f'B{row}'] = breakdown.get('non_compliant', 0)
    
    def create_compliance_matrix_sheet(self, matrix_data):
        """Create compliance matrix sheet"""
        
        ws = self.workbook.create_sheet("Compliance Matrix")
        
        # Title
        ws['A1'] = matrix_data.get('title', 'Compliance Matrix')
        ws['A1'].font = Font(size=16, bold=True)
        
        # Headers
        headers = matrix_data.get('headers', [
            "Clause ID", "Category", "Parameter", "Required", 
            "Meter Spec", "Status", "Justification", "Risk", "Comments"
        ])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        data_rows = matrix_data.get('data', [])
        for row_idx, row_data in enumerate(data_rows, 4):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_meter_specs_sheet(self, specs_data):
        """Create meter specifications sheet"""
        
        ws = self.workbook.create_sheet("Meter Specifications")
        
        # Title
        ws['A1'] = specs_data.get('title', 'Meter Specifications')
        ws['A1'].font = Font(size=16, bold=True)
        
        # Meter details
        meter_details = specs_data.get('meter_details', {})
        
        row = 3
        ws[f'A{row}'] = "Model:"
        ws[f'B{row}'] = meter_details.get('model', 'Unknown')
        
        row += 1
        ws[f'A{row}'] = "Series:"
        ws[f'B{row}'] = meter_details.get('series', 'Unknown')
        
        row += 1
        ws[f'A{row}'] = "Selection Source:"
        ws[f'B{row}'] = meter_details.get('selection_source', 'Unknown')
        
        # Specifications
        row += 2
        ws[f'A{row}'] = "Specifications:"
        ws[f'A{row}'].font = Font(bold=True)
        
        specifications = meter_details.get('specifications', {})
        for spec_key, spec_value in specifications.items():
            row += 1
            ws[f'A{row}'] = f"{spec_key.replace('_', ' ').title()}:"
            ws[f'B{row}'] = str(spec_value)